"""
Knight Insurance — Document Processor
Handles text extraction from all document types:
  - Digital PDFs (direct text extraction via PyMuPDF)
  - Scanned PDFs / photo PDFs (OCR via pytesseract)
  - Images (OCR)
  - Excel files (structured parsing)
  - Word documents (text extraction)
"""
import os
import io
import json
import logging
from typing import Tuple, Optional
from pathlib import Path

from PIL import Image

logger = logging.getLogger(__name__)

# Prefer PyMuPDF (fitz) for PDF extraction — much better table handling
try:
    import fitz  # PyMuPDF
    FITZ_AVAILABLE = True
except ImportError:
    FITZ_AVAILABLE = False
    logger.warning("PyMuPDF not available, falling back to PyPDF2")

try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

# Optional imports — graceful degradation if not installed
try:
    import pytesseract
    from pdf2image import convert_from_path
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    logger.warning("pytesseract or pdf2image not available. OCR disabled.")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False


class DocumentProcessor:
    """Extracts text and structured data from various document formats."""

    # Minimum chars per page to consider PDF as "digital" (not scanned)
    MIN_CHARS_PER_PAGE = 50

    def process(self, file_path: str) -> Tuple[str, Optional[dict], float]:
        """
        Process a document and extract its content.

        Returns:
            Tuple of (extracted_text, structured_data, quality_score)
            - extracted_text: raw text content
            - structured_data: dict for Excel/structured docs, None for text docs
            - quality_score: 0.0 to 1.0 (1.0 = perfect digital, 0.5 = OCR, etc.)
        """
        file_ext = Path(file_path).suffix.lower()

        try:
            if file_ext == ".pdf":
                return self._process_pdf(file_path)
            elif file_ext in (".xlsx", ".xls"):
                return self._process_excel(file_path)
            elif file_ext == ".docx":
                return self._process_docx(file_path)
            elif file_ext in (".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".gif"):
                return self._process_image(file_path)
            elif file_ext == ".csv":
                return self._process_csv(file_path)
            else:
                # Try to read as plain text
                return self._process_text(file_path)
        except Exception as e:
            logger.error(f"Error processing {file_path}: {e}")
            return f"[Error processing file: {e}]", None, 0.0

    def _process_pdf(self, file_path: str) -> Tuple[str, None, float]:
        """Extract text from PDF using PyMuPDF (fitz) for superior table handling."""
        if FITZ_AVAILABLE:
            return self._process_pdf_fitz(file_path)
        elif PYPDF2_AVAILABLE:
            return self._process_pdf_pypdf2(file_path)
        else:
            return "[No PDF library available]", None, 0.0

    def _process_pdf_fitz(self, file_path: str) -> Tuple[str, None, float]:
        """Extract text using PyMuPDF — best for complex tables and layouts."""
        doc = fitz.open(file_path)
        pages_text = []
        total_chars = 0

        for page in doc:
            text = page.get_text("text")
            pages_text.append(text)
            total_chars += len(text.strip())

        doc.close()
        num_pages = len(pages_text) or 1
        chars_per_page = total_chars / num_pages

        if chars_per_page >= self.MIN_CHARS_PER_PAGE:
            full_text = "\n\n--- Page Break ---\n\n".join(pages_text)
            quality = min(1.0, chars_per_page / 500)
            logger.info(f"PyMuPDF extracted {total_chars} chars from {num_pages} pages ({chars_per_page:.0f} chars/page)")
            return full_text, None, quality

        # Likely scanned — try OCR
        if OCR_AVAILABLE:
            logger.info(f"PDF has low text content ({chars_per_page:.0f} chars/page). Attempting OCR...")
            return self._ocr_pdf(file_path)
        else:
            logger.warning("OCR not available. Returning limited text from scanned PDF.")
            return "\n\n".join(pages_text), None, 0.2

    def _process_pdf_pypdf2(self, file_path: str) -> Tuple[str, None, float]:
        """Fallback: Extract text using PyPDF2."""
        reader = PyPDF2.PdfReader(file_path)
        pages_text = []
        total_chars = 0

        for page in reader.pages:
            text = page.extract_text() or ""
            pages_text.append(text)
            total_chars += len(text.strip())

        num_pages = len(reader.pages) or 1
        chars_per_page = total_chars / num_pages

        if chars_per_page >= self.MIN_CHARS_PER_PAGE:
            full_text = "\n\n--- Page Break ---\n\n".join(pages_text)
            quality = min(1.0, chars_per_page / 500)
            return full_text, None, quality

        if OCR_AVAILABLE:
            logger.info(f"PDF has low text content ({chars_per_page:.0f} chars/page). Attempting OCR...")
            return self._ocr_pdf(file_path)
        else:
            logger.warning("OCR not available. Returning limited text from scanned PDF.")
            return "\n\n".join(pages_text), None, 0.2

    def _ocr_pdf(self, file_path: str) -> Tuple[str, None, float]:
        """OCR a scanned PDF using pdf2image + pytesseract."""
        try:
            images = convert_from_path(file_path, dpi=300)
            pages_text = []
            for i, img in enumerate(images):
                text = pytesseract.image_to_string(img)
                pages_text.append(text)
                logger.debug(f"OCR page {i+1}: {len(text)} chars")

            full_text = "\n\n--- Page Break ---\n\n".join(pages_text)
            # OCR quality is inherently lower
            quality = 0.6 if len(full_text.strip()) > 100 else 0.3
            return full_text, None, quality
        except Exception as e:
            logger.error(f"OCR failed for {file_path}: {e}")
            return f"[OCR failed: {e}]", None, 0.1

    def _process_excel(self, file_path: str) -> Tuple[str, dict, float]:
        """Parse Excel file into structured data."""
        if not EXCEL_AVAILABLE:
            return "[Excel parsing not available]", None, 0.0

        wb = openpyxl.load_workbook(file_path, data_only=True)
        structured = {}
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            headers = None

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                # Skip completely empty rows
                values = [str(c) if c is not None else "" for c in row]
                if all(v == "" for v in values):
                    continue

                if headers is None and any(v != "" for v in values):
                    # Check if this looks like a header row (has multiple non-empty cells)
                    non_empty = sum(1 for v in values if v != "")
                    if non_empty >= 2:
                        headers = values
                        text_parts.append(f"Sheet: {sheet_name}")
                        text_parts.append(" | ".join(headers))
                        continue

                if headers:
                    row_dict = {}
                    for i, val in enumerate(values):
                        if i < len(headers) and headers[i]:
                            row_dict[headers[i]] = val
                    if any(v != "" for v in row_dict.values()):
                        rows.append(row_dict)
                        text_parts.append(" | ".join(values))
                else:
                    # Pre-header row (like title)
                    non_empty_vals = [v for v in values if v]
                    if non_empty_vals:
                        text_parts.append(" ".join(non_empty_vals))

            structured[sheet_name] = {
                "headers": headers,
                "rows": rows,
                "row_count": len(rows)
            }

        full_text = "\n".join(text_parts)
        return full_text, structured, 1.0

    def _process_docx(self, file_path: str) -> Tuple[str, None, float]:
        """Extract text from Word document."""
        if not DOCX_AVAILABLE:
            return "[DOCX parsing not available]", None, 0.0

        doc = docx.Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        full_text = "\n".join(paragraphs)
        return full_text, None, 1.0

    def _process_image(self, file_path: str) -> Tuple[str, None, float]:
        """Extract text from image using Gemini Vision API."""
        try:
            from utils.gemini import call_gemini_vision

            with open(file_path, "rb") as f:
                image_data = f.read()

            ext = Path(file_path).suffix.lower()
            mime_map = {
                ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                ".gif": "image/gif", ".bmp": "image/bmp", ".tiff": "image/tiff",
                ".webp": "image/webp",
            }
            mime_type = mime_map.get(ext, "image/png")

            prompt = (
                "Extract ALL text and data from this image. "
                "If this is a driver's license or CDL, extract: "
                "full name, date of birth, license number, state, class, "
                "endorsements, restrictions, expiration date, sex, height, weight, eye color. "
                "Format the output as structured text with clear labels."
            )

            text = call_gemini_vision(
                prompt, image_data, mime_type=mime_type,
                step_name="image_ocr"
            )
            quality = 0.85 if len(text.strip()) > 50 else 0.5
            logger.info(f"Gemini Vision extracted {len(text)} chars from {file_path}")
            return text, None, quality
        except Exception as e:
            logger.error(f"Gemini Vision image processing failed: {e}")
            # Fallback to OCR if available
            if OCR_AVAILABLE:
                try:
                    img = Image.open(file_path)
                    text = pytesseract.image_to_string(img)
                    return text, None, 0.6
                except Exception:
                    pass
            return f"[Image processing failed: {e}]", None, 0.1

    def _process_csv(self, file_path: str) -> Tuple[str, dict, float]:
        """Read CSV file."""
        import csv
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            return "[Empty CSV]", None, 0.5

        headers = rows[0]
        data_rows = []
        for row in rows[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val
            data_rows.append(row_dict)

        structured = {"headers": headers, "rows": data_rows, "row_count": len(data_rows)}
        text = "\n".join([" | ".join(row) for row in rows])
        return text, structured, 1.0

    def _process_text(self, file_path: str) -> Tuple[str, None, float]:
        """Read plain text file."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                text = f.read()
            return text, None, 1.0
        except Exception as e:
            return f"[Error reading text file: {e}]", None, 0.0

    def get_file_type(self, filename: str) -> str:
        """Determine file type from extension."""
        ext = Path(filename).suffix.lower()
        type_map = {
            ".pdf": "pdf",
            ".xlsx": "xlsx",
            ".xls": "xls",
            ".csv": "csv",
            ".docx": "docx",
            ".doc": "doc",
            ".png": "image",
            ".jpg": "image",
            ".jpeg": "image",
            ".tiff": "image",
            ".bmp": "image",
            ".gif": "image",
            ".txt": "text",
        }
        return type_map.get(ext, "unknown")
