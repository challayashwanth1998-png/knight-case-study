"""
Generate perfect sample submission files (PDF + Excel) that pass ALL underwriting rules.
Files are designed with clear, explicit field labels so the AI extractor can find all data.
"""
import os
import sys

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_DIR = os.path.join(os.path.dirname(__file__), "perfect-submission")
os.makedirs(OUT_DIR, exist_ok=True)

NAVY = colors.HexColor("#1E3A8A")
LIGHT_BLUE = colors.HexColor("#EFF6FF")
GREEN_BG = colors.HexColor("#DCFCE7")


def _styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle('T', parent=styles['Title'], fontSize=16, textColor=NAVY, spaceAfter=4),
        "sub": ParagraphStyle('S', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER),
        "heading": ParagraphStyle('H', parent=styles['Heading2'], fontSize=12, textColor=NAVY, spaceBefore=14, spaceAfter=4),
        "normal": styles['Normal'],
        "note": ParagraphStyle('N', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor("#166534")),
    }


def _info_table(data, col1=2*inch, col2=4.5*inch):
    t = Table(data, colWidths=[col1, col2])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    return t


def _data_table(data, widths=None):
    if widths is None:
        widths = [1.3*inch] * len(data[0])
    t = Table(data, colWidths=widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    return t


# ════════════════════════════════════════════════════════════
# 1. APPLICATION PDF
# ════════════════════════════════════════════════════════════
def create_application_pdf():
    path = os.path.join(OUT_DIR, "Heartland_Express_Application.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    s = _styles()
    e = []

    e.append(Paragraph("COMMERCIAL AUTO LIABILITY APPLICATION", s["title"]))
    e.append(Paragraph("Knight Specialty Insurance Program", s["sub"]))
    e.append(Spacer(1, 10))

    # ── Applicant ──
    e.append(Paragraph("APPLICANT INFORMATION", s["heading"]))
    e.append(_info_table([
        ["Company Name:", "Heartland Express Trucking LLC"],
        ["DBA:", "Heartland Express"],
        ["Business Type:", "LLC"],
        ["FEIN:", "42-1987654"],
        ["DOT Number:", "2847291"],
        ["MC Number:", "MC-891234"],
        ["Owner Name:", "James R. Mitchell"],
        ["Date Established:", "March 15, 2012"],
        ["Years in Business:", "14"],
        ["Years Trucking Exp:", "18"],
    ]))

    # ── Address ──
    e.append(Paragraph("MAILING ADDRESS", s["heading"]))
    e.append(_info_table([
        ["Street:", "1250 Industrial Blvd, Suite 200"],
        ["City:", "Nashville"],
        ["State:", "TN"],
        ["Zip:", "37210"],
    ]))

    # ── Operations ──
    e.append(Paragraph("OPERATIONS", s["heading"]))
    e.append(_info_table([
        ["Operation Type:", "For-hire long-haul general freight"],
        ["Interstate:", "Yes"],
        ["Intrastate:", "No"],
        ["Radius:", "500+ miles (Nationwide)"],
        ["Commodities Hauled:", "General freight, dry goods, consumer electronics, packaged food products"],
        ["Annual Revenue:", "$4,200,000"],
        ["Annual Mileage:", "1,850,000"],
        ["Hazmat:", "No"],
        ["Hazmat Details:", "N/A - No hazardous materials transported"],
        ["Towing/Recovery:", "No"],
        ["Mexico Border:", "No"],
        ["Intermodal/Container:", "No"],
    ]))

    # ── Fleet ──
    e.append(Paragraph("FLEET INFORMATION", s["heading"]))
    e.append(_info_table([
        ["Total Power Units:", "5 (all semi-trucks)"],
        ["Total Trailers:", "4"],
        ["Regular Drivers:", "5"],
        ["Part-Time Drivers:", "0"],
        ["Owner Operators:", "0"],
        ["Total Drivers:", "5"],
    ]))

    # ── Coverage ──
    e.append(Paragraph("COVERAGE REQUESTED", s["heading"]))
    e.append(_info_table([
        ["Auto Liability:", "$1,000,000 CSL"],
        ["Physical Damage:", "Not Requested - Not Applicable"],
        ["Cargo Coverage:", "Yes - $100,000"],
        ["General Liability:", "No"],
        ["Medical Payments:", "Yes - $5,000"],
        ["Uninsured Motorist:", "$1,000,000"],
        ["Estimated Annual Premium:", "$75,000"],
    ]))

    # ── Loss History ──
    e.append(Paragraph("LOSS HISTORY (3 YEARS)", s["heading"]))
    e.append(_data_table([
        ["Year", "# Claims", "Total Incurred", "Details"],
        ["2024", "0", "$0", "No claims"],
        ["2023", "1", "$12,500", "Minor fender bender, not at-fault"],
        ["2022", "0", "$0", "No claims"],
    ], [1*inch, 1*inch, 1.2*inch, 3.3*inch]))
    e.append(Spacer(1, 4))
    e.append(Paragraph("Total 3-Year Losses: $12,500 | Loss Ratio: 2.1%", s["note"]))

    # ── Prior Insurance ──
    e.append(Paragraph("PRIOR INSURANCE", s["heading"]))
    e.append(_info_table([
        ["Current Carrier:", "National Indemnity"],
        ["Policy Number:", "NAT-2024-TRK-88912"],
        ["Effective Date:", "September 1, 2024"],
        ["Expiration Date:", "September 1, 2026"],
        ["Current Premium:", "$185,000"],
        ["Cancellations:", "None"],
        ["Non-Renewals:", "None"],
    ]))

    # ── Safety ──
    e.append(Paragraph("SAFETY & COMPLIANCE", s["heading"]))
    e.append(_info_table([
        ["Drug Testing:", "Yes — random quarterly testing"],
        ["Safety Program:", "Monthly safety meetings, dash cameras, ELD compliant"],
        ["Hiring Standards:", "2+ years CDL experience, clean MVR, current DOT physical"],
    ]))

    doc.build(e)
    print(f"✅ Created: {path}")


# ════════════════════════════════════════════════════════════
# 2. DRIVER ROSTER (Excel)
# ════════════════════════════════════════════════════════════
def create_driver_roster_xlsx():
    path = os.path.join(OUT_DIR, "Heartland_Express_Driver_Roster.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Driver Roster"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    cen = Alignment(horizontal='center')

    ws.merge_cells('A1:L1')
    ws['A1'] = "HEARTLAND EXPRESS TRUCKING LLC — DRIVER ROSTER"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws['A2'] = "As of June 15, 2026 — 5 Active Drivers"
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    headers = ["#", "Name", "Date of Birth", "Age", "CDL Number", "CDL State",
               "CDL Class", "Years Experience", "Hire Date", "Violations 3yr", "Accidents 3yr", "Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = cen; cell.border = bd

    drivers = [
        [1, "James Mitchell",    "1978-04-12", 48, "TN-88234591", "TN", "A", 18, "2015-06-01", "None", "None", "Active"],
        [2, "Robert Williams",   "1985-09-22", 40, "TN-77123488", "TN", "A", 12, "2018-03-15", "None", "None", "Active"],
        [3, "David Thompson",    "1990-01-30", 36, "AL-55987234", "AL", "A",  8, "2020-01-10", "None", "None", "Active"],
        [4, "Michael Garcia",    "1982-07-15", 43, "AR-44321876", "AR", "A", 15, "2016-09-01", "None", "None", "Active"],
        [5, "Kevin Brown",       "1988-11-05", 37, "TN-99876543", "TN", "A", 10, "2019-07-20", "None", "None", "Active"],
    ]

    gf = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    for ri, drv in enumerate(drivers, 5):
        for ci, val in enumerate(drv, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bd; cell.alignment = cen
            if ci in (10, 11): cell.fill = gf

    widths = [4, 20, 13, 6, 15, 10, 10, 14, 12, 14, 14, 8]
    for i, w in enumerate(widths, 1): ws.column_dimensions[chr(64+i)].width = w

    summary_row = 5 + len(drivers) + 1
    ws.merge_cells(f'A{summary_row}:L{summary_row}')
    c = ws.cell(row=summary_row, column=1, value="SUMMARY: All 5 drivers meet requirements — Age 23+ ✓ | 2+ Years CDL ✓ | Valid CDL ✓ | Clean MVR ✓ | No Violations ✓ | No Accidents ✓")
    c.font = Font(bold=True, color="166534", size=10); c.fill = gf

    wb.save(path)
    print(f"✅ Created: {path}")


# ════════════════════════════════════════════════════════════
# 3. VEHICLE SCHEDULE (Excel)
# ════════════════════════════════════════════════════════════
def create_vehicle_schedule_xlsx():
    path = os.path.join(OUT_DIR, "Heartland_Express_Vehicle_Schedule.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Schedule"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    cen = Alignment(horizontal='center')

    ws.merge_cells('A1:G1')
    ws['A1'] = "HEARTLAND EXPRESS TRUCKING LLC — VEHICLE / EQUIPMENT SCHEDULE"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws['A2'] = "5 Power Units (All Semi-Trucks) + 4 Trailers"
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    ws.merge_cells('A3:G3')
    ws['A3'] = "POWER UNITS"
    ws['A3'].font = Font(bold=True, size=12, color="1E3A8A")

    headers = ["Unit #", "Year", "Make / Model", "VIN", "Vehicle Type", "GVW", "Stated Value"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = cen; cell.border = bd

    vehicles = [
        [1, 2022, "Freightliner Cascadia", "3AKJHHDR5NSLA1234", "Semi-Truck (Sleeper)", "80,000 lbs", "$125,000"],
        [2, 2021, "Kenworth T680",         "1XKYD49X1MJ456789", "Semi-Truck (Sleeper)", "80,000 lbs", "$115,000"],
        [3, 2023, "Peterbilt 579",         "2NP2HN0X8PT234567", "Semi-Truck (Day Cab)", "80,000 lbs", "$140,000"],
        [4, 2022, "Volvo VNL 860",         "4V4NC9EH3NN345678", "Semi-Truck (Sleeper)", "80,000 lbs", "$130,000"],
        [5, 2020, "International LT",      "3HSDJAPR5LN567890", "Semi-Truck (Day Cab)", "80,000 lbs", "$95,000"],
    ]
    for ri, v in enumerate(vehicles, 5):
        for ci, val in enumerate(v, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bd; cell.alignment = cen

    trailer_start = 5 + len(vehicles) + 1
    ws.merge_cells(f'A{trailer_start}:G{trailer_start}')
    ws.cell(row=trailer_start, column=1, value="TRAILERS").font = Font(bold=True, size=12, color="1E3A8A")

    header_row = trailer_start + 1
    for c, h in enumerate(["#", "Year", "Make / Model", "VIN", "Type", "Length", ""], 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = cen; cell.border = bd

    trailers = [
        [1, 2021, "Great Dane Everest",  "1GRAA0622NB123456", "Dry Van Trailer", "53 ft", ""],
        [2, 2022, "Utility 4000D-X",     "1UYVS2530NM234567", "Dry Van Trailer", "53 ft", ""],
        [3, 2020, "Wabash National",     "1JJV532B8NL345678", "Dry Van Trailer", "53 ft", ""],
        [4, 2022, "Great Dane Champion",  "1GRAA8826NN678901", "Reefer Trailer",  "53 ft", ""],
    ]
    data_start = header_row + 1
    for ri, t in enumerate(trailers, data_start):
        for ci, val in enumerate(t, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bd; cell.alignment = cen

    widths = [8, 6, 22, 20, 22, 12, 12]
    for i, w in enumerate(widths, 1): ws.column_dimensions[chr(64+i)].width = w

    wb.save(path)
    print(f"✅ Created: {path}")


# ════════════════════════════════════════════════════════════
# 4. IFTA REPORTS — 4 QUARTERS (one PDF with all 4)
# ════════════════════════════════════════════════════════════
def create_ifta_reports():
    """Create 4 separate IFTA PDFs (one per quarter) so the system counts 4 IFTA reports."""
    # Only allowed states: AL, AZ, AR, CA, SC, SD, TN, TX, UT, VA, WV, WI, WY
    quarters = [
        ("Q1 2026", "January - March 2026"),
        ("Q2 2025", "April - June 2025"),
        ("Q3 2025", "July - September 2025"),
        ("Q4 2025", "October - December 2025"),
    ]
    
    states_data = [
        ["State", "Miles", "Gallons", "Tax Paid"],
        ["TN", "72,000", "10,800", "$2,160.00"],
        ["AL", "48,000", "7,200", "$1,200.00"],
        ["AR", "36,000", "5,400", "$1,188.00"],
        ["VA", "42,000", "6,300", "$1,134.00"],
        ["SC", "30,000", "4,500", "$810.00"],
        ["WI", "22,000", "3,300", "$1,056.00"],
        ["AZ", "19,000", "2,850", "$541.50"],
        ["UT", "14,000", "2,100", "$640.50"],
    ]

    for q_name, q_period in quarters:
        fname = f"Heartland_Express_IFTA_{q_name.replace(' ', '_')}.pdf"
        path = os.path.join(OUT_DIR, fname)
        doc = SimpleDocTemplate(path, pagesize=letter, topMargin=0.5*inch)
        s = _styles()
        e = []

        e.append(Paragraph("INTERNATIONAL FUEL TAX AGREEMENT (IFTA)", s["title"]))
        e.append(Paragraph(f"Quarterly Report — {q_name} ({q_period})", s["sub"]))
        e.append(Spacer(1, 10))

        e.append(Paragraph("CARRIER INFORMATION", s["heading"]))
        e.append(_info_table([
            ["Carrier Name:", "Heartland Express Trucking LLC"],
            ["IFTA License:", "TN-4829-1876"],
            ["DOT Number:", "2847291"],
            ["MC Number:", "MC-891234"],
            ["FEIN:", "42-1987654"],
            ["Base State:", "TN"],
            ["Quarter:", q_name],
            ["Period:", q_period],
        ]))

        e.append(Paragraph("JURISDICTIONAL MILES & FUEL PURCHASES", s["heading"]))
        total_miles = sum(int(r[1].replace(",","")) for r in states_data[1:])
        total_gal = sum(int(r[2].replace(",","")) for r in states_data[1:])
        total_tax = sum(float(r[3].replace("$","").replace(",","")) for r in states_data[1:])
        full_data = states_data + [["TOTAL", f"{total_miles:,}", f"{total_gal:,}", f"${total_tax:,.2f}"]]
        
        t = _data_table(full_data, [1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), LIGHT_BLUE),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        e.append(t)

        e.append(Spacer(1, 8))
        states_list = ", ".join(r[0] for r in states_data[1:])
        e.append(Paragraph(
            f"States Traveled: {states_list} — All states are within Knight's approved territory.",
            s["note"]
        ))

        doc.build(e)
        print(f"✅ Created: {path}")


# ════════════════════════════════════════════════════════════
# 5. LOSS RUNS (PDF)
# ════════════════════════════════════════════════════════════
def create_loss_runs_pdf():
    path = os.path.join(OUT_DIR, "Heartland_Express_Loss_Runs_3Year.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter, topMargin=0.5*inch)
    s = _styles()
    e = []

    e.append(Paragraph("LOSS RUNS — 3 YEAR HISTORY", s["title"]))
    e.append(Paragraph("National Indemnity Insurance | Policy: NAT-2024-TRK-88912", s["sub"]))
    e.append(Spacer(1, 10))

    e.append(Paragraph("INSURED INFORMATION", s["heading"]))
    e.append(_info_table([
        ["Insured:", "Heartland Express Trucking LLC"],
        ["FEIN:", "42-1987654"],
        ["DOT Number:", "2847291"],
        ["MC Number:", "MC-891234"],
        ["Valuation Date:", "June 15, 2026"],
    ]))

    e.append(Paragraph("POLICY YEAR SUMMARY", s["heading"]))
    summary = [
        ["Policy Year", "Earned Premium", "# Claims", "Total Incurred", "Total Paid", "Loss Ratio"],
        ["2025-2026", "$185,000", "0", "$0", "$0", "0.0%"],
        ["2024-2025", "$185,000", "0", "$0", "$0", "0.0%"],
        ["2023-2024", "$178,000", "1", "$12,500", "$12,500", "7.0%"],
        ["2022-2023", "$172,000", "0", "$0", "$0", "0.0%"],
        ["3-Year Total", "$535,000", "1", "$12,500", "$12,500", "2.3%"],
    ]
    t = Table(summary, colWidths=[1.3*inch, 1.2*inch, 0.8*inch, 1.1*inch, 1*inch, 1*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), GREEN_BG),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    e.append(t)

    e.append(Paragraph("CLAIM DETAIL", s["heading"]))
    claim = [
        ["Claim #", "Date of Loss", "Driver", "Description", "Status", "Incurred"],
        ["CLM-2023-0891", "2023-08-14", "Steven Anderson",
         "Rear-ended at stoplight by third party. Not at-fault. Minor bumper damage only.", "Closed", "$12,500"],
    ]
    t = Table(claim, colWidths=[1.2*inch, 0.9*inch, 1.1*inch, 2.2*inch, 0.7*inch, 0.8*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    e.append(t)

    e.append(Spacer(1, 16))
    e.append(Paragraph("3-Year Loss Ratio: 2.3% — Excellent loss history. One not-at-fault claim in 3 years.", s["note"]))

    doc.build(e)
    print(f"✅ Created: {path}")


# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Generating perfect submission sample files...\n")
    create_application_pdf()
    create_driver_roster_xlsx()
    create_vehicle_schedule_xlsx()
    create_ifta_reports()       # 4 separate PDFs
    create_loss_runs_pdf()
    print(f"\n📁 All files saved to: {OUT_DIR}")
    print("📧 Attach ALL files to an email and send to sarah@outreachbenefits.online")
    print("   → 1 Application PDF")
    print("   → 1 Driver Roster XLSX")
    print("   → 1 Vehicle Schedule XLSX")
    print("   → 4 IFTA PDFs (one per quarter)")
    print("   → 1 Loss Runs PDF")
    print("   → Total: 8 files")
