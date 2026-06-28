"""
Generate complete sample data for ABC Freight Systems.
Creates all documents needed for a full underwriting submission.
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

OUTPUT = "/Users/challa/Documents/knight-insurance/sample-data/complete-submission"

# ─── 1. INSURANCE APPLICATION (TXT) ──────────────────────────
def create_application():
    content = """
═══════════════════════════════════════════════════════════════
              COMMERCIAL TRUCK FLEET APPLICATION
              Knight Specialty Insurance Program
═══════════════════════════════════════════════════════════════

SECTION 1: GENERAL INFORMATION
─────────────────────────────────────────────────────────────
Business Name:           ABC Freight Systems, Inc.
DBA:                     ABC Freight
Business Type:           Corporation
FEIN:                    82-4917365
MC Number:               MC-847291
DOT Number:              3284716
Date Business Started:   03/15/2018
Years in Trucking:       7

SECTION 2: MAILING ADDRESS
─────────────────────────────────────────────────────────────
Street:                  4521 Industrial Blvd, Suite 200
City:                    Dallas
State:                   TX
ZIP:                     75247

SECTION 3: GARAGING ADDRESS
─────────────────────────────────────────────────────────────
Street:                  4521 Industrial Blvd
City:                    Dallas
State:                   TX
ZIP:                     75247

SECTION 4: OWNER / PRINCIPAL
─────────────────────────────────────────────────────────────
Owner Name:              Robert M. Chen
Owner SSN:               XXX-XX-4829
Owner DOB:               06/12/1975
Owner Phone:             (214) 555-0187
Owner Email:             rchen@abcfreight.com

SECTION 5: DESCRIPTION OF OPERATIONS
─────────────────────────────────────────────────────────────
Operation Type:          For-Hire Carrier
Interstate:              Yes
Intrastate:              Yes
Radius of Operations:    Regional / Long Haul (500+ miles)
Commodities Hauled:      General Freight, Dry Goods, Consumer Electronics,
                         Building Materials, Packaged Foods
Hazmat:                  No
Hazmat Details:          N/A

SECTION 6: FLEET INFORMATION
─────────────────────────────────────────────────────────────
Number of Power Units:   12
Number of Trailers:      15
Total Vehicles:          27

SECTION 7: DRIVER INFORMATION
─────────────────────────────────────────────────────────────
Regular Drivers:         10
Part-Time Drivers:       2
Owner Operators:         0
Total Drivers:           12

SECTION 8: COVERAGE REQUESTED
─────────────────────────────────────────────────────────────
Auto Liability Limit:    $1,000,000 CSL
Physical Damage:         No
Cargo Coverage:          Yes
Cargo Limit:             $100,000
General Liability:       No
Medical Payments:        No
Uninsured Motorist:      No
Filing Required:         Yes (BMC-91)

SECTION 9: PRIOR INSURANCE
─────────────────────────────────────────────────────────────
Current Carrier:         National Interstate Insurance
Policy Number:           NI-TRK-2024-08847
Policy Period:           07/01/2024 - 07/01/2025
Premium:                 $187,500
Cancelled/Non-Renewed:   No
Cancellation Reason:     N/A

SECTION 10: SAFETY MEASURES
─────────────────────────────────────────────────────────────
Employment Background Check:     Yes
Pre-Employment Drug Test:        Yes
Criminal Background Check:       Yes
MVR Review (Pre-Employment):     Yes
Road Test Required:              Yes
PSP Report Review:               Yes
Annual MVR Review:               Yes
Random Drug Testing:             Yes
Telematics/GPS:                  Yes (Samsara)
Dash Cameras:                    Yes (Forward + Driver-Facing)
Safety Director:                 Yes (Maria Santos)
Written Safety Program:          Yes
Smith System Training:           Yes

SECTION 11: STATES OF OPERATION
─────────────────────────────────────────────────────────────
Primary: TX, AR, TN, AL, VA
Secondary: AZ, CA, UT, WI, SC

SECTION 12: METROPOLITAN AREAS
─────────────────────────────────────────────────────────────
Dallas-Fort Worth, Houston, San Antonio, Memphis, Nashville,
Little Rock, Birmingham, Phoenix

SECTION 13: MULTIPLE TERMINALS
─────────────────────────────────────────────────────────────
Terminal 1 (Primary):    Dallas, TX
Terminal 2:              Memphis, TN

═══════════════════════════════════════════════════════════════
Applicant Signature: Robert M. Chen     Date: 06/15/2026
═══════════════════════════════════════════════════════════════
"""
    with open(os.path.join(OUTPUT, "ABC_Freight_Insurance_Application.txt"), "w") as f:
        f.write(content)
    print("✓ Insurance Application")


# ─── 2. DETAILED DRIVER ROSTER (XLSX) ────────────────────────
def create_driver_roster():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Driver Roster"

    # Header
    header_font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A1"] = "ABC Freight Systems, Inc. — Driver Roster"
    ws["A1"].font = header_font
    ws.merge_cells("A2:J2")
    ws["A2"] = "As of June 15, 2026 | MC-847291 | DOT-3284716"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    # Column headers
    headers = [
        "Driver #", "Full Name", "Date of Birth", "Age",
        "CDL Number", "CDL State", "CDL Class",
        "Date of Hire", "Years CDL Experience", "Status"
    ]
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_ft = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_ft
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    drivers = [
        [1, "Robert M. Chen", "06/12/1975", 51, "TX-82749163", "TX", "A", "03/15/2018", 18, "Active - Owner"],
        [2, "James A. Williams", "09/23/1982", 43, "TX-63841927", "TX", "A", "06/01/2019", 14, "Active"],
        [3, "Carlos R. Martinez", "03/07/1988", 38, "TX-51739284", "TX", "A", "01/15/2020", 10, "Active"],
        [4, "David L. Thompson", "11/18/1979", 46, "AR-47281936", "AR", "A", "04/22/2020", 16, "Active"],
        [5, "Michael J. Davis", "07/30/1990", 35, "TX-38194726", "TX", "A", "08/10/2020", 8, "Active"],
        [6, "William K. Anderson", "02/14/1985", 41, "TN-62937481", "TN", "A", "11/03/2021", 12, "Active"],
        [7, "Brian S. Jackson", "12/05/1992", 33, "TX-29471638", "TX", "A", "03/18/2022", 6, "Active"],
        [8, "Anthony P. Garcia", "04/21/1987", 39, "TX-83641972", "TX", "A", "07/25/2022", 11, "Active"],
        [9, "Kevin R. Wilson", "08/16/1983", 42, "AL-57293814", "AL", "A", "01/09/2023", 13, "Active"],
        [10, "Steven T. Moore", "10/29/1991", 34, "TX-41628379", "TX", "A", "05/14/2023", 7, "Active"],
        [11, "Daniel E. Taylor", "05/03/1994", 32, "VA-73849216", "VA", "A", "09/01/2024", 5, "Active - Part Time"],
        [12, "Christopher L. Brown", "01/17/1986", 40, "TX-64917283", "TX", "A", "02/15/2025", 12, "Active - Part Time"],
    ]

    for row_idx, driver in enumerate(drivers, 5):
        for col_idx, val in enumerate(driver, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")

    # Summary
    ws.cell(row=18, column=1, value="Total Active Drivers:").font = Font(bold=True)
    ws.cell(row=18, column=2, value=12)
    ws.cell(row=19, column=1, value="Average Age:").font = Font(bold=True)
    ws.cell(row=19, column=2, value=39.5)
    ws.cell(row=20, column=1, value="Average CDL Experience (years):").font = Font(bold=True)
    ws.cell(row=20, column=2, value=11.0)
    ws.cell(row=21, column=1, value="All drivers age 23+:").font = Font(bold=True)
    ws.cell(row=21, column=2, value="Yes")
    ws.cell(row=22, column=1, value="All drivers 2+ years CDL:").font = Font(bold=True)
    ws.cell(row=22, column=2, value="Yes")

    # Column widths
    for col, w in zip("ABCDEFGHIJ", [10, 25, 14, 6, 16, 10, 10, 14, 20, 18]):
        ws.column_dimensions[col].width = w

    wb.save(os.path.join(OUTPUT, "ABC_Freight_Driver_Roster.xlsx"))
    print("✓ Driver Roster")


# ─── 3. DETAILED EQUIPMENT SCHEDULE (XLSX) ───────────────────
def create_equipment_schedule():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment Schedule"

    ws.merge_cells("A1:H1")
    ws["A1"] = "ABC Freight Systems, Inc. — Equipment Schedule"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:H2")
    ws["A2"] = "As of June 15, 2026 | MC-847291 | DOT-3284716"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    headers = ["Unit #", "Year", "Make", "Model", "VIN", "Vehicle Type", "GVW", "Ownership"]
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_ft = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_ft
        cell.fill = header_fill

    vehicles = [
        # Power Units (Semi-Trucks)
        ["PU-001", 2023, "Freightliner", "Cascadia 126", "3AKJHHDR5PSNA7841", "Semi-Truck (Sleeper)", "80,000", "Owned"],
        ["PU-002", 2023, "Freightliner", "Cascadia 126", "3AKJHHDR7PSNA7842", "Semi-Truck (Sleeper)", "80,000", "Owned"],
        ["PU-003", 2022, "Kenworth", "T680", "1XKYD49X7NJ148293", "Semi-Truck (Sleeper)", "80,000", "Owned"],
        ["PU-004", 2022, "Kenworth", "T680", "1XKYD49X9NJ148294", "Semi-Truck (Sleeper)", "80,000", "Owned"],
        ["PU-005", 2022, "Peterbilt", "579", "1XPBD49X2ND847291", "Semi-Truck (Sleeper)", "80,000", "Owned"],
        ["PU-006", 2021, "Volvo", "VNL 860", "4V4NC9EH8MN281637", "Semi-Truck (Sleeper)", "80,000", "Owned"],
        ["PU-007", 2021, "Volvo", "VNL 860", "4V4NC9EH0MN281638", "Semi-Truck (Sleeper)", "80,000", "Owned"],
        ["PU-008", 2021, "Freightliner", "Cascadia 116", "3AKJGLDR3LSJA9127", "Semi-Truck (Day Cab)", "80,000", "Owned"],
        ["PU-009", 2020, "Kenworth", "T680", "1XKYD49X4LJ283946", "Semi-Truck (Day Cab)", "80,000", "Leased"],
        ["PU-010", 2020, "Peterbilt", "579", "1XPBD49X6LD738291", "Semi-Truck (Sleeper)", "80,000", "Leased"],
        ["PU-011", 2019, "Freightliner", "Cascadia 126", "3AKJHHDR8KSLA8374", "Semi-Truck (Sleeper)", "80,000", "Leased"],
        ["PU-012", 2023, "Kenworth", "T680", "1XKYD49X1PJ384729", "Semi-Truck (Day Cab)", "80,000", "Owned"],
        # Trailers
        ["TR-001", 2023, "Great Dane", "Champion CL", "1GRAA0622PB847291", "Dry Van Trailer", "—", "Owned"],
        ["TR-002", 2023, "Great Dane", "Champion CL", "1GRAA0624PB847292", "Dry Van Trailer", "—", "Owned"],
        ["TR-003", 2022, "Utility", "4000D-X", "1UYVS2539NU384721", "Dry Van Trailer", "—", "Owned"],
        ["TR-004", 2022, "Utility", "4000D-X", "1UYVS2531NU384722", "Dry Van Trailer", "—", "Owned"],
        ["TR-005", 2022, "Wabash", "DuraPlate", "1JJV532D9NL847293", "Dry Van Trailer", "—", "Owned"],
        ["TR-006", 2021, "Great Dane", "Champion CL", "1GRAA0628MB384726", "Dry Van Trailer", "—", "Owned"],
        ["TR-007", 2021, "Utility", "4000D-X", "1UYVS2533MU284927", "Dry Van Trailer", "—", "Owned"],
        ["TR-008", 2021, "Wabash", "DuraPlate", "1JJV532D1ML847298", "Dry Van Trailer", "—", "Leased"],
        ["TR-009", 2020, "Great Dane", "Everest SS", "1GRAA0620LB293847", "Reefer Trailer", "—", "Leased"],
        ["TR-010", 2020, "Utility", "3000R", "1UYVS2535LU847210", "Reefer Trailer", "—", "Leased"],
        ["TR-011", 2023, "Fontaine", "Revolution", "13N14820XP1384711", "Flatbed Trailer", "—", "Owned"],
        ["TR-012", 2022, "Fontaine", "Infinity", "13N14820XN1384712", "Flatbed Trailer", "—", "Owned"],
        ["TR-013", 2021, "East", "Genesis", "1E1U5Y286ML384713", "Flatbed Trailer", "—", "Owned"],
        ["TR-014", 2023, "Great Dane", "Champion CL", "1GRAA0626PB847214", "Dry Van Trailer", "—", "Owned"],
        ["TR-015", 2022, "Wabash", "DuraPlate", "1JJV532D3NL847215", "Dry Van Trailer", "—", "Owned"],
    ]

    for row_idx, v in enumerate(vehicles, 5):
        for col_idx, val in enumerate(v, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    r = 5 + len(vehicles) + 1
    ws.cell(row=r, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=r+1, column=1, value="Total Power Units (Semi-Trucks):").font = Font(bold=True)
    ws.cell(row=r+1, column=2, value=12)
    ws.cell(row=r+2, column=1, value="Total Trailers:").font = Font(bold=True)
    ws.cell(row=r+2, column=2, value=15)
    ws.cell(row=r+3, column=1, value="Total Fleet:").font = Font(bold=True)
    ws.cell(row=r+3, column=2, value=27)
    ws.cell(row=r+4, column=1, value="All Semi-Trucks:").font = Font(bold=True)
    ws.cell(row=r+4, column=2, value="Yes (100% eligible)")
    ws.cell(row=r+5, column=1, value="Avg Vehicle Age:").font = Font(bold=True)
    ws.cell(row=r+5, column=2, value="3.2 years")

    for col, w in zip("ABCDEFGH", [10, 8, 16, 18, 22, 24, 10, 12]):
        ws.column_dimensions[col].width = w

    wb.save(os.path.join(OUTPUT, "ABC_Freight_Equipment_Schedule.xlsx"))
    print("✓ Equipment Schedule")


# ─── 4. LOSS RUNS — CURRENT + 3 PRIOR YEARS (TXT) ───────────
def create_loss_runs():
    content = """
═══════════════════════════════════════════════════════════════════
                     LOSS RUN REPORT
                National Interstate Insurance
═══════════════════════════════════════════════════════════════════
Insured:           ABC Freight Systems, Inc.
Policy Number:     NI-TRK-2024-08847
Valuation Date:    06/10/2026
Report Generated:  06/12/2026

═══════════════════════════════════════════════════════════════════
CURRENT POLICY PERIOD: 07/01/2024 – 07/01/2025
═══════════════════════════════════════════════════════════════════

Claim #     Date of Loss  Description                    Status   Incurred    Paid     Reserve
─────────── ────────────  ─────────────────────────────  ──────── ─────────── ──────── ────────
NI-24-00147 09/14/2024    Rear-end collision, I-30       Closed   $12,450     $12,450  $0
                          Dallas TX. Minor damage.
                          At-fault: No

NI-24-00203 12/02/2024    Side-swipe on I-40 Memphis     Closed   $8,720      $8,720   $0
                          TN. Property damage only.
                          At-fault: Yes (Driver #6)

Period Total: 2 claims | Total Incurred: $21,170 | Total Paid: $21,170

═══════════════════════════════════════════════════════════════════
PRIOR YEAR 1: 07/01/2023 – 07/01/2024
Policy: NI-TRK-2023-07293
═══════════════════════════════════════════════════════════════════

Claim #     Date of Loss  Description                    Status   Incurred    Paid     Reserve
─────────── ────────────  ─────────────────────────────  ──────── ─────────── ──────── ────────
NI-23-00089 10/22/2023    Cargo shift during braking.    Closed   $4,200      $4,200   $0
                          Minor cargo damage claim.
                          At-fault: No

Period Total: 1 claim | Total Incurred: $4,200 | Total Paid: $4,200

═══════════════════════════════════════════════════════════════════
PRIOR YEAR 2: 07/01/2022 – 07/01/2023
Policy: NI-TRK-2022-06184
═══════════════════════════════════════════════════════════════════

Claim #     Date of Loss  Description                    Status   Incurred    Paid     Reserve
─────────── ────────────  ─────────────────────────────  ──────── ─────────── ──────── ────────
NI-22-00312 02/15/2023    Weather-related accident       Closed   $18,900     $18,900  $0
                          I-20 near Abilene TX.
                          Bodily injury third party.
                          At-fault: Partial

NI-22-00044 09/08/2022    Parking lot backing incident.  Closed   $3,100      $3,100   $0
                          Property damage only.
                          At-fault: Yes (Driver #3)

Period Total: 2 claims | Total Incurred: $22,000 | Total Paid: $22,000

═══════════════════════════════════════════════════════════════════
PRIOR YEAR 3: 07/01/2021 – 07/01/2022
Policy: NI-TRK-2021-05092
═══════════════════════════════════════════════════════════════════

No claims during this period.

Period Total: 0 claims | Total Incurred: $0 | Total Paid: $0

═══════════════════════════════════════════════════════════════════
                        4-YEAR SUMMARY
═══════════════════════════════════════════════════════════════════
Total Claims:        5
Total Incurred:      $47,370
Total Paid:          $47,370
Open Reserves:       $0
Loss Ratio:          6.3% (based on $750,000 4-year premium)
At-Fault Claims:     2 of 5 (40%)
Bodily Injury:       1 of 5 (20%)

Comments: Generally favorable loss history. No catastrophic losses.
Average claim size: $9,474. Two at-fault incidents in 4 years
across 12 drivers represents acceptable frequency.
═══════════════════════════════════════════════════════════════════
"""
    with open(os.path.join(OUTPUT, "ABC_Freight_Loss_Runs_4Year.txt"), "w") as f:
        f.write(content)
    print("✓ Loss Runs (Current + 3 Prior Years)")


# ─── 5. 4th IFTA QUARTER (Q4 2025) ──────────────────────────
def create_ifta_q4():
    content = """
═══════════════════════════════════════════════════════════════
           IFTA QUARTERLY TAX RETURN
           4th Quarter 2025 (October - December)
═══════════════════════════════════════════════════════════════
Company:     ABC Freight Systems, Inc.
Address:     4521 Industrial Blvd, Suite 200, Dallas, TX 75247
IFTA License: TX-847291-IFTA
Fleet MPG:   6.42

STATE    TOTAL MILES   TAX GALLONS   TAX PAID GAL   NET TAX GAL   TAX RATE    TAX DUE
──────   ──────────   ───────────   ────────────   ───────────   ─────────   ────────
TX        48,291        7,522          8,100         -578         $0.2000    ($115.60)
AR        18,472        2,877          2,400          477         $0.2450     $116.87
TN        15,893        2,476          2,200          276         $0.2700      $74.52
AL        12,384        1,929          1,800          129         $0.2900      $37.41
AZ         8,729        1,360          1,100          260         $0.2600      $67.60
VA         6,482        1,010            900          110         $0.2620      $28.82
SC         4,218          657            500          157         $0.2800      $43.96
UT         3,891          606            400          206         $0.3195      $65.82
CA         2,847          443            300          143         $0.6830      $97.67
WI         1,293          201            100          101         $0.3290      $33.23
──────   ──────────   ───────────   ────────────   ───────────               ────────
TOTAL    122,500       19,081         17,800        1,281                     $450.30

SURCHARGES:
  IN Surface Transportation — 122,500 x 0.0000 = $0.00

GRAND TOTAL IFTA TAX: $450.30
States Traveled: TX, AR, TN, AL, AZ, VA, SC, UT, CA, WI (10 states)
═══════════════════════════════════════════════════════════════
"""
    with open(os.path.join(OUTPUT, "ABC_Freight_IFTA_4Q_2025.txt"), "w") as f:
        f.write(content)
    print("✓ IFTA Q4 2025")


# ─── 6. MVRs FOR ALL DRIVERS (XLSX) ─────────────────────────
def create_mvrs():
    wb = openpyxl.Workbook()

    drivers_mvr = [
        ("Robert M. Chen", "TX-82749163", "TX", "06/12/1975", "A", "Valid", [], [], 0),
        ("James A. Williams", "TX-63841927", "TX", "09/23/1982", "A", "Valid",
         [("03/15/2024", "Speeding 10 mph over", 2, True)], [], 2),
        ("Carlos R. Martinez", "TX-51739284", "TX", "03/07/1988", "A", "Valid",
         [("11/02/2023", "Failure to signal lane change", 1, True)], [], 1),
        ("David L. Thompson", "AR-47281936", "AR", "11/18/1979", "A", "Valid", [], [], 0),
        ("Michael J. Davis", "TX-38194726", "TX", "07/30/1990", "A", "Valid",
         [("08/20/2024", "Speeding 15 mph over", 3, True)], [], 3),
        ("William K. Anderson", "TN-62937481", "TN", "02/14/1985", "A", "Valid", [], [], 0),
        ("Brian S. Jackson", "TX-29471638", "TX", "12/05/1992", "A", "Valid", [],
         [("05/10/2024", "Minor fender bender parking lot", False)], 0),
        ("Anthony P. Garcia", "TX-83641972", "TX", "04/21/1987", "A", "Valid", [], [], 0),
        ("Kevin R. Wilson", "AL-57293814", "AL", "08/16/1983", "A", "Valid",
         [("01/30/2025", "Speeding 8 mph over", 1, True)], [], 1),
        ("Steven T. Moore", "TX-41628379", "TX", "10/29/1991", "A", "Valid", [], [], 0),
        ("Daniel E. Taylor", "VA-73849216", "VA", "05/03/1994", "A", "Valid", [], [], 0),
        ("Christopher L. Brown", "TX-64917283", "TX", "01/17/1986", "A", "Valid", [], [], 0),
    ]

    ws = wb.active
    ws.title = "MVR Summary"

    ws.merge_cells("A1:K1")
    ws["A1"] = "ABC Freight Systems — Motor Vehicle Records Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:K2")
    ws["A2"] = "Records pulled: June 10, 2026 | Source: State DMV databases"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    headers = ["Driver Name", "CDL Number", "State", "DOB", "CDL Class",
               "License Status", "Violations (3yr)", "Points (3yr)",
               "Points (12mo)", "Accidents (3yr)", "DUI/DWI"]
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for i, d in enumerate(drivers_mvr, 5):
        name, lic, st, dob, cls, status, violations, accidents, pts = d
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=lic)
        ws.cell(row=i, column=3, value=st)
        ws.cell(row=i, column=4, value=dob)
        ws.cell(row=i, column=5, value=cls)
        ws.cell(row=i, column=6, value=status)
        ws.cell(row=i, column=7, value=len(violations))
        ws.cell(row=i, column=8, value=pts)
        pts_12 = sum(v[2] for v in violations if "2025" in v[0] or "2026" in v[0])
        ws.cell(row=i, column=9, value=pts_12)
        ws.cell(row=i, column=10, value=len(accidents))
        ws.cell(row=i, column=11, value="No")

    r = 5 + len(drivers_mvr) + 1
    ws.cell(row=r, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=r+1, column=1, value="All licenses valid:").font = Font(bold=True)
    ws.cell(row=r+1, column=2, value="Yes")
    ws.cell(row=r+2, column=1, value="DUI/DWI/Reckless:").font = Font(bold=True)
    ws.cell(row=r+2, column=2, value="None")
    ws.cell(row=r+3, column=1, value="Max points (3yr):").font = Font(bold=True)
    ws.cell(row=r+3, column=2, value="3 (Michael J. Davis)")
    ws.cell(row=r+4, column=1, value="Max points (12mo):").font = Font(bold=True)
    ws.cell(row=r+4, column=2, value="1 (Kevin R. Wilson)")
    ws.cell(row=r+5, column=1, value="Drivers exceeding 6pts/3yr:").font = Font(bold=True)
    ws.cell(row=r+5, column=2, value="None")
    ws.cell(row=r+6, column=1, value="Drivers exceeding 4pts/12mo:").font = Font(bold=True)
    ws.cell(row=r+6, column=2, value="None")

    for col, w in zip("ABCDEFGHIJK", [24, 18, 8, 12, 10, 14, 16, 14, 14, 14, 10]):
        ws.column_dimensions[col].width = w

    # Individual MVR sheets
    for d in drivers_mvr:
        name, lic, st, dob, cls, status, violations, accidents, pts = d
        safe = name.replace(" ", "_").replace(".", "")[:28]
        mvr_ws = wb.create_sheet(title=safe)
        mvr_ws["A1"] = "MOTOR VEHICLE RECORD"
        mvr_ws["A1"].font = Font(bold=True, size=14)
        mvr_ws["A3"] = "Driver Name:"
        mvr_ws["B3"] = name
        mvr_ws["A4"] = "CDL Number:"
        mvr_ws["B4"] = lic
        mvr_ws["A5"] = "State:"
        mvr_ws["B5"] = st
        mvr_ws["A6"] = "Date of Birth:"
        mvr_ws["B6"] = dob
        mvr_ws["A7"] = "CDL Class:"
        mvr_ws["B7"] = cls
        mvr_ws["A8"] = "License Status:"
        mvr_ws["B8"] = status
        mvr_ws["A9"] = "DUI/DWI:"
        mvr_ws["B9"] = "No"
        mvr_ws["A10"] = "Reckless Driving:"
        mvr_ws["B10"] = "No"
        mvr_ws["A11"] = "Hit and Run:"
        mvr_ws["B11"] = "No"
        mvr_ws["A12"] = "Felony:"
        mvr_ws["B12"] = "No"

        mvr_ws["A14"] = "VIOLATIONS (Last 3 Years)"
        mvr_ws["A14"].font = Font(bold=True)
        if violations:
            mvr_ws["A15"] = "Date"
            mvr_ws["B15"] = "Description"
            mvr_ws["C15"] = "Points"
            mvr_ws["D15"] = "Conviction"
            for j, v in enumerate(violations, 16):
                mvr_ws.cell(row=j, column=1, value=v[0])
                mvr_ws.cell(row=j, column=2, value=v[1])
                mvr_ws.cell(row=j, column=3, value=v[2])
                mvr_ws.cell(row=j, column=4, value="Yes" if v[3] else "No")
        else:
            mvr_ws["A15"] = "No violations in the last 3 years."

        r2 = 17 + len(violations)
        mvr_ws.cell(row=r2, column=1, value="ACCIDENTS (Last 3 Years)").font = Font(bold=True)
        if accidents:
            for j, a in enumerate(accidents, r2 + 1):
                mvr_ws.cell(row=j, column=1, value=a[0])
                mvr_ws.cell(row=j, column=2, value=a[1])
                mvr_ws.cell(row=j, column=3, value="At-Fault" if a[2] else "Not At-Fault")
        else:
            mvr_ws.cell(row=r2 + 1, column=1, value="No accidents in the last 3 years.")

        mvr_ws.column_dimensions["A"].width = 20
        mvr_ws.column_dimensions["B"].width = 40
        mvr_ws.column_dimensions["C"].width = 10
        mvr_ws.column_dimensions["D"].width = 12

    wb.save(os.path.join(OUTPUT, "ABC_Freight_MVR_All_Drivers.xlsx"))
    print("✓ MVRs (All 12 Drivers)")


# ─── RUN ALL ─────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"Generating files in: {OUTPUT}\n")
    create_application()
    create_driver_roster()
    create_equipment_schedule()
    create_loss_runs()
    create_ifta_q4()
    create_mvrs()

    # Copy existing IFTA reports
    import shutil
    src = "/Users/challa/Documents/knight-insurance/sample-data"
    for f in os.listdir(src):
        if "IFTA" in f and f.endswith(".pdf"):
            shutil.copy2(os.path.join(src, f), os.path.join(OUTPUT, f))
            print(f"✓ Copied {f}")

    print(f"\n✅ All files generated! Upload everything from:\n   {OUTPUT}")
