"""Quick regeneration of just Application PDF + Driver Roster + Vehicle Schedule."""
import os
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(__file__), "perfect-submission")
NAVY = colors.HexColor("#1E3A8A")

styles = getSampleStyleSheet()
S = {
    "title": ParagraphStyle('T', parent=styles['Title'], fontSize=16, textColor=NAVY, spaceAfter=4),
    "sub": ParagraphStyle('S', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER),
    "heading": ParagraphStyle('H', parent=styles['Heading2'], fontSize=12, textColor=NAVY, spaceBefore=14, spaceAfter=4),
    "normal": styles['Normal'],
    "note": ParagraphStyle('N', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor("#166534")),
}

def _t(data, c1=2*inch, c2=4.5*inch):
    t = Table(data, colWidths=[c1, c2])
    t.setStyle(TableStyle([('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('BOTTOMPADDING',(0,0),(-1,-1),3),('TOPPADDING',(0,0),(-1,-1),3)]))
    return t

def _dt(data, widths=None):
    if not widths: widths = [1.3*inch]*len(data[0])
    t = Table(data, colWidths=widths)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,colors.grey),('ALIGN',(1,1),(-1,-1),'RIGHT'),('BOTTOMPADDING',(0,0),(-1,-1),4),('TOPPADDING',(0,0),(-1,-1),4)]))
    return t

# 1. APPLICATION PDF
print("Creating application PDF...")
path = os.path.join(OUT, "Heartland_Express_Application.pdf")
doc = SimpleDocTemplate(path, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
e = []
e.append(Paragraph("COMMERCIAL AUTO LIABILITY APPLICATION", S["title"]))
e.append(Paragraph("Knight Specialty Insurance Program", S["sub"]))
e.append(Spacer(1, 10))

e.append(Paragraph("APPLICANT INFORMATION", S["heading"]))
e.append(_t([
    ["Company Name:", "Heartland Express Trucking LLC"],
    ["DBA:", "Heartland Express"],
    ["Business Type:", "LLC"],
    ["FEIN:", "42-1987654"],
    ["DOT Number:", "2847291"],
    ["MC Number:", "MC-891234"],
    ["Owner Name:", "James R. Mitchell"],
    ["Date Established:", "March 15, 2012"],
    ["Years in Business:", "14"],
    ["Years Trucking Exp:", "18"],
]))

e.append(Paragraph("MAILING ADDRESS", S["heading"]))
e.append(_t([
    ["Street:", "1250 Industrial Blvd, Suite 200"],
    ["City:", "Nashville"], ["State:", "TN"], ["Zip:", "37210"],
]))

e.append(Paragraph("OPERATIONS", S["heading"]))
e.append(_t([
    ["Operation Type:", "For-hire long-haul general freight"],
    ["Interstate:", "Yes"], ["Intrastate:", "No"],
    ["Radius:", "500+ miles (Nationwide)"],
    ["Commodities Hauled:", "General freight, dry goods, consumer electronics, packaged food products"],
    ["Annual Revenue:", "$4,200,000"],
    ["Annual Mileage:", "1,850,000"],
    ["Hazmat:", "No"],
    ["Towing/Recovery:", "No"],
    ["Mexico Border:", "No"],
    ["Intermodal/Container:", "No"],
]))

e.append(Paragraph("FLEET INFORMATION", S["heading"]))
e.append(_t([
    ["Total Power Units:", "5 (all semi-trucks)"],
    ["Total Trailers:", "4"],
    ["Regular Drivers:", "5"],
    ["Part-Time Drivers:", "0"],
    ["Owner Operators:", "0"],
    ["Total Drivers:", "5"],
]))

e.append(Paragraph("COVERAGE REQUESTED", S["heading"]))
e.append(_t([
    ["Auto Liability:", "$1,000,000 CSL"],
    ["Physical Damage:", "Yes - Comprehensive & Collision, ACV basis"],
    ["Cargo Coverage:", "Yes - $100,000"],
    ["General Liability:", "No"],
    ["Medical Payments:", "Yes - $5,000"],
    ["Uninsured Motorist:", "$1,000,000"],
    ["Estimated Annual Premium:", "$75,000"],
]))

e.append(Paragraph("LOSS HISTORY (3 YEARS)", S["heading"]))
e.append(_dt([
    ["Year", "# Claims", "Total Incurred", "Details"],
    ["2024", "0", "$0", "No claims"],
    ["2023", "1", "$12,500", "Minor fender bender, not at-fault"],
    ["2022", "0", "$0", "No claims"],
], [1*inch, 1*inch, 1.2*inch, 3.3*inch]))
e.append(Spacer(1, 4))
e.append(Paragraph("Total 3-Year Losses: $12,500 | Loss Ratio: 2.1%", S["note"]))

e.append(Paragraph("PRIOR INSURANCE", S["heading"]))
e.append(_t([
    ["Current Carrier:", "National Indemnity"],
    ["Policy Number:", "NAT-2024-TRK-88912"],
    ["Effective Date:", "September 1, 2024"],
    ["Expiration Date:", "September 1, 2026"],
    ["Current Premium:", "$185,000"],
    ["Cancelled or Non-Renewed:", "No - continuous coverage since 2018"],
    ["Cancellation Reason:", "N/A"],
]))

e.append(Paragraph("SAFETY & COMPLIANCE", S["heading"]))
e.append(_t([
    ["Employment Background Check:", "Yes"],
    ["Pre-Employment Drug Test:", "Yes - DOT 5-panel"],
    ["Criminal Background Check:", "Yes"],
    ["MVR Review (Pre-Employment):", "Yes"],
    ["Annual MVR Review:", "Yes"],
    ["Road Test:", "Yes - all new hires"],
    ["Telematics / GPS:", "Yes - Samsara fleet tracking on all units"],
    ["Safety Director:", "Yes - James Mitchell (Owner/Safety Director)"],
    ["Drug Testing Program:", "Yes - random quarterly DOT testing"],
    ["Safety Meetings:", "Yes - monthly, documented"],
    ["Dash Cameras:", "Yes - forward and driver-facing on all trucks"],
    ["ELD Compliance:", "Yes - fully compliant"],
]))

doc.build(e)
print(f"  OK: {path}")

# 2. DRIVER ROSTER
print("Creating driver roster...")
path = os.path.join(OUT, "Heartland_Express_Driver_Roster.xlsx")
wb = Workbook(); ws = wb.active; ws.title = "Driver Roster"
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
cen = Alignment(horizontal='center')

ws.merge_cells('A1:L1')
ws['A1'] = "HEARTLAND EXPRESS TRUCKING LLC - DRIVER ROSTER"
ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
ws['A2'] = "As of June 15, 2026 - 5 Active Drivers"
ws['A2'].font = Font(italic=True, size=10, color="666666")

headers = ["#", "Name", "Date of Birth", "Age", "CDL Number", "CDL State", "CDL Class", "Years Experience", "Hire Date", "Violations 3yr", "Accidents 3yr", "Status"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.alignment = cen; cell.border = bd

drivers = [
    [1, "James Mitchell",    "1978-04-12", 48, "TN-88234591", "TN", "A", 18, "2015-06-01", 0, 0, "Active"],
    [2, "Robert Williams",   "1985-09-22", 40, "TN-77123488", "TN", "A", 12, "2018-03-15", 0, 0, "Active"],
    [3, "David Thompson",    "1990-01-30", 36, "AL-55987234", "AL", "A",  8, "2020-01-10", 0, 0, "Active"],
    [4, "Michael Garcia",    "1982-07-15", 43, "AR-44321876", "AR", "A", 15, "2016-09-01", 0, 0, "Active"],
    [5, "Kevin Brown",       "1988-11-05", 37, "TN-99876543", "TN", "A", 10, "2019-07-20", 0, 0, "Active"],
]
gf = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
for ri, drv in enumerate(drivers, 5):
    for ci, val in enumerate(drv, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border = bd; cell.alignment = cen
        if ci in (10, 11): cell.fill = gf

widths = [4, 20, 13, 6, 15, 10, 10, 14, 12, 14, 14, 8]
for i, w in enumerate(widths, 1): ws.column_dimensions[chr(64+i)].width = w

sr = 5 + len(drivers) + 1
ws.merge_cells(f'A{sr}:L{sr}')
c = ws.cell(row=sr, column=1, value="SUMMARY: All 5 drivers meet requirements - Age 23+ | 2+ Years CDL | Valid CDL | Clean MVR | 0 Violations | 0 Accidents")
c.font = Font(bold=True, color="166534", size=10); c.fill = gf
wb.save(path)
print(f"  OK: {path}")

# 3. VEHICLE SCHEDULE
print("Creating vehicle schedule...")
path = os.path.join(OUT, "Heartland_Express_Vehicle_Schedule.xlsx")
wb = Workbook(); ws = wb.active; ws.title = "Vehicle Schedule"

ws.merge_cells('A1:G1')
ws['A1'] = "HEARTLAND EXPRESS TRUCKING LLC - VEHICLE / EQUIPMENT SCHEDULE"
ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
ws['A2'] = "5 Power Units (All Semi-Trucks) + 4 Trailers"
ws['A2'].font = Font(italic=True, size=10, color="666666")

ws.merge_cells('A3:G3')
ws['A3'] = "POWER UNITS"
ws['A3'].font = Font(bold=True, size=12, color="1E3A8A")

headers = ["Unit #", "Year", "Make / Model", "VIN", "Vehicle Type", "GVW", "Stated Value"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.alignment = cen; cell.border = bd

vehicles = [
    ["T-101", 2022, "Freightliner Cascadia", "3AKJHHDR5NSLA1234", "Semi-Truck (Sleeper)", "80,000 lbs", "$125,000"],
    ["T-102", 2021, "Kenworth T680",         "1XKYD49X1MJ456789", "Semi-Truck (Sleeper)", "80,000 lbs", "$115,000"],
    ["T-103", 2023, "Peterbilt 579",         "2NP2HN0X8PT234567", "Semi-Truck (Day Cab)", "80,000 lbs", "$140,000"],
    ["T-104", 2022, "Volvo VNL 860",         "4V4NC9EH3NN345678", "Semi-Truck (Sleeper)", "80,000 lbs", "$130,000"],
    ["T-105", 2020, "International LT",      "3HSDJAPR5LN567890", "Semi-Truck (Day Cab)", "80,000 lbs", "$95,000"],
]
for ri, v in enumerate(vehicles, 5):
    for ci, val in enumerate(v, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border = bd; cell.alignment = cen

ts = 5 + len(vehicles) + 1
ws.merge_cells(f'A{ts}:G{ts}')
ws.cell(row=ts, column=1, value="TRAILERS").font = Font(bold=True, size=12, color="1E3A8A")

hr = ts + 1
for c, h in enumerate(["#", "Year", "Make / Model", "VIN", "Type", "Length", ""], 1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.alignment = cen; cell.border = bd

trailers = [
    ["TR-201", 2021, "Great Dane Everest", "1GRAA0622MB123456", "Dry Van", "53 ft"],
    ["TR-202", 2022, "Utility 4000D-X",    "1UYVS2530NM654321", "Dry Van", "53 ft"],
    ["TR-203", 2020, "Wabash National DuraPlate", "1JJV532B5ML987654", "Dry Van", "53 ft"],
    ["TR-204", 2023, "Hyundai Translead",  "3H3V532C6PT246813", "Reefer", "53 ft"],
]
for ri, tr in enumerate(trailers, hr + 1):
    for ci, val in enumerate(tr, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border = bd; cell.alignment = cen

for i, w in enumerate([10, 8, 22, 22, 18, 12, 12], 1):
    ws.column_dimensions[chr(64+i)].width = w

wb.save(path)
print(f"  OK: {path}")
print("\nAll 3 files regenerated!")
